import time                                                                   # İşlemler arası bekleme yapmak için kullanacağımız sleep() methodunun barındırıldığı kütüphane
from selenium import webdriver                                                # Proje konumuz olan ana kütüphane
from selenium.webdriver.common.by import By                                   # Scrapper için yardımcı kütüphane
from selenium.webdriver.common.keys import Keys                               # Scrapper için yardımcı kütüphane 
from openpyxl import Workbook, load_workbook                                  # Excell'e veri kaydetmek içi kütüphane


wb = Workbook()                                                               # Excell çalışma dosyası tanımlıyoruz
ws = wb.active                                                                # Dosyamızı aktifleştiriyoruz
ws.title = "Twit-Data"                                                        # Dosyamıza başlık verdik


                                                     ##### ÖNEMLİ #####
driver_path = "C:/Users/Bad/Desktop/chromedriver"    ##### ÖNEMLİ #####  =>   # chromedriver'ın yolunuzu belirtmelisiniz, bende masaüstünde olduğu için bu şekilde belirttim
                                                     ##### ÖNEMLİ #####
try:                                                                          # Scrapping'de beklenmedik hatalar olabiliyor, programın yarıda kalmaması için try finally bloğu kullanıyoruz
    Qut = 0
    while Qut == 0:                                                           # Kullanıcının programı tekrar başlatmadan çeşitli işlemler yapabilmesi için döngüye alıyoruz
        print("İşlem Seçiniz")
        print("--------------")
        print("1: #Hastag'e göre twitlerin çekilmesi")  
        print("2: @Kullanıcı twitleri")                 
        print("3: @Kullanıcı - ♥Like'ları")             
        print("0: Çıkış")
        secim = int(input("--------------" + "\n"))
    
        if secim < 0 or secim > 3:                                            # Kullanıcı doğru bir seçim mi yapmış kontrolü
            print("Hatalı işlem seçtiniz")
    
        elif secim == 1:
    
            
            #EXCEL#
            ws = wb.create_sheet("Hastag Tweets")                             # Hastag Tweets adında excell sayfası oluşturuyoruz
            ws.append(["Kullanıcı Adı", "Tweet"])                             # Excell tablo başlıklarını ekliyoruz
            #######
    
    
            #İstenilen HASTAG
            hashtag = input("Twitlerini istediğiniz #hastag:" + "\n")         # arama yapılacak #hastag'in alınması
            
            #Browser açılması
            browser = webdriver.Chrome(executable_path=driver_path)           # tarayıcımız açan kod
            browser.get("https://twitter.com/login")                          # tarayıcıda twitter giriş sayfasına gidiyoruz
            time.sleep(3)                                                     # hata almamak için sıradaki işleme kadar sayfanın yüklenmesini bekliyoruz
    
            #Giriş
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[1]/label/div/div[2]/div/input").send_keys("botum04539141") # giriş ekranında kullanıcı adımızı giriyoruz
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[2]/label/div/div[2]/div/input").send_keys("yenibot07")     # giriş ekranında şifremizi giriyoruz
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[3]/div/div").click()                                       # giriş tuşuna basıyoruz. sayfada bu elemanlardan birer tane olduğu için xpath kullanıyoruz
            
            #Arama
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div[2]/div/div[2]/div/div/div/div[1]/div/div/div/form/div[1]/div/div/div[2]/input").send_keys(hashtag)    # arama çubuğuna aranılacak #hastagi yazıyoruz
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div[2]/div/div[2]/div/div/div/div[1]/div/div/div/form/div[1]/div/div/div[2]/input").send_keys(Keys.ENTER) # aramaya başlamak için ENTER tuşuna bastırıyoruz
            time.sleep(5)                                                                                                                                                        # hata almamak için sıradaki işleme kadar sayfanın yüklenmesini bekliyoruz
            
            
            kullanıcı = []                              # kullanı adlarını tutacağımız liste
            tweets = []                                 # tweetleri tutacağımız liste
            
            #Twitlerin çekilmesi
            for i in range(15):                         # 15 semboliktir. twitterin ne kadar sonuç göstereceği değişebiliyor. 15 kere sayfayı aşağıya kaydırıyoruz genelde tüm hastaglari çekmek için yeterli oluyor
            
                kadi = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[1]/div/div/div[1]/div[1]/a/div/div[2]")    # şuandaki sayfanın tüm article taglarının içerinde belirlenen yolu izleyerek kullanıcı adlarını çekiyoruz
                for i in kadi:
                    kullaniciadi = i.text
                    kullaniciadisafe = kullaniciadi.replace("@", "")                                                     # Excell kaydında hata almamak için kullanıcı adlarının başındaki "@"" işaretini kırpıyoruz
                    kullanıcı.append(kullaniciadisafe.strip())                                                           # Kullanıcı adını son alarak boşluklardanda kurtararak temize çekiyoruz
                twit = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[1]")                                # şuandaki sayfanın tüm article taglarının içerinde belirlenen yolu izleyerek tweetleri çekiyoruz
                for i in twit:
                    tw = i.text
                    tweets.append(tw.strip())                                                                            # Yukarıdaki işlemin benzerleri
                browser.find_element_by_tag_name('html').send_keys(Keys.END)                                             # sayfadaki tweetlerle işi bitince react yüklemesi için sayfayı aşağıya kaydırıyoruz ve yeni tweetler ekleniyor
            
                
            
            count = len(kullanıcı)                                                                                       # çekilen tweet adedimiz (kullanıcı adı baz alınır)
            print("Tweet sayısı: " + str(count))
            
            #Excel kayıt
            for i in range(0, count):                                                                                    # tweetlerimizi Excell'e aktarmak için döngüye alıyoruz
                ws.append([kullanıcı[i], tweets[i]])                                                                     # kayıtlar excell tablosuna ekleniyor, buradan sonra bir sıkıntı çıkmayacağı için her döngüde kayıt almak yerine..
            wb.save("twit-data.xlsx")                                                                                    # tüm kayıtlar eklenince Excell tablomuzu kaydediyoruz
    
        elif secim == 2:
            
            #EXCEL#
            ws = wb.create_sheet("User tweets")
            ws.append(["Tweet", "Reply", "Retweet", "Like"])
            #######
    
            #Kullanıcı adı
            print("--------------")
            username = input("Twitlerini istediğiniz @Kullanıcı:" + "\n")
            print("--------------")
            twitcount = int(input("İstenilen en az twit sayısı:" +"\n"))
            
    
            #Kullanıcı giriş kontrolü
            while twitcount <= 0:
                print("--------------")
                print("1 den az olamaz")
                twitcount = int(input("İstenilen en az twit sayısı:" +"\n"))
    
            
            #Browser açılması
            browser = webdriver.Chrome(executable_path=driver_path)
            browser.get("https://twitter.com/login")
            time.sleep(3)
    
            #Giriş
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[1]/label/div/div[2]/div/input").send_keys("botum04539141")
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[2]/label/div/div[2]/div/input").send_keys("yenibot07")
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[3]/div/div").click()
            
            #Profile yönlendirme
            profileurl = "https://twitter.com/" + str(username)
            browser.get(profileurl)
    
            tweets = []
            replies = []
            retweets = []
            likes = []
    
            SCROLL_PAUSE_TIME = 1
            last_height = browser.execute_script("return document.body.scrollHeight")          
            
            while True:
    
                #Twitler
                twit = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[1]")
                for i in twit:
                    tw = i.text
                    tweets.append(tw.strip())
                    
                #Replyler
                reply = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[3]/div[1]")
                for i in reply:
                    rp = i.text
                    replies.append(rp.strip())
                
                #Retweetler
                retweet = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[3]/div[2]")
                for i in retweet:
                    rt = i.text
                    retweets.append(rt.strip())
                
                #Likes
                like = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[3]/div[3]")
                for i in like:
                    lk = i.text
                    likes.append(lk.strip())
    
               
                #İstenen twit çekildimi kontrolü
                counted = len(tweets)
                if counted >= twitcount:
                    break
    
                # Ekranı aşağı kaydırma
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                # Yenileme beklmesi
                time.sleep(SCROLL_PAUSE_TIME)
    
                
                # Ekran en aşağıda mı kontrolü 
                new_height = browser.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            
            count = len(tweets)
            print("Tweet sayısı: " + str(count))
            
            #Excel kayıt
            for i in range(0, count):
                ws.append([tweets[i], replies[i],retweets[i],likes[i]])
            wb.save("user-twit-data.xlsx")
            browser.quit()
        
        elif secim == 3:
            #EXCEL#
            ws = wb.create_sheet("User likes")
            ws.append(["User","Tweet"])
            #######
    
            #Kullanıcı adı
            print("--------------")
            username = input("♥ Likelarını istediğiniz @Kullanıcı:" + "\n")
            print("--------------")
            twitcount = int(input("İstenilen en az like sayısı:" +"\n"))
            
    
            #Kullanıcı giriş kontrolü
            while twitcount <= 0:
                print("--------------")
                print("1 den az olamaz")
                twitcount = int(input("İstenilen en az like sayısı:" +"\n"))
    
            #Browser açılması
            browser = webdriver.Chrome(executable_path=driver_path)
            browser.get("https://twitter.com/login")
            time.sleep(3)
    
            #Giriş
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[1]/label/div/div[2]/div/input").send_keys("botum04539141")
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[2]/label/div/div[2]/div/input").send_keys("yenibot07")
            browser.find_element_by_xpath(
                "//*[@id='react-root']/div/div/div[2]/main/div/div/div[1]/form/div/div[3]/div/div").click()
            
            #Profile yönlendirme
            profileurl = "https://twitter.com/" + str(username)
            browser.get(profileurl)
    
            time.sleep(3)
            browser.find_element_by_xpath("//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[2]/div/div/nav/div/div[2]/div/div[4]/a").click()
    
            users = []
            tweets = []
            replies = []
            retweets = []
            likes = []
    
            SCROLL_PAUSE_TIME = 1
            last_height = browser.execute_script("return document.body.scrollHeight")
            
            while True:
                #Likelar
    
                user = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[1]/div/div/div[1]/div[1]/a/div/div[2]")
                for i in user:
                    us_forsafe = i.text
                    us = us_forsafe.replace("@", "")
                    users.append(us.strip())
                
                twit = browser.find_elements_by_xpath(
                    "//article[@role='article']/div/div/div/div[2]/div[2]/div[2]/div[1]")
                for i in twit:
                    if i.text:
                        tw = i.text
                        tweets.append(tw.strip())
                    else:
                        replies.append("null")
                
    
               
                #İstenen twit çekildimi kontrolü
                counted = len(users)
                if counted >= twitcount:
                    break
    
                # Ekranı aşağı kaydırma
                browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                # Yenileme beklmesi
                time.sleep(SCROLL_PAUSE_TIME)
    
                
                # Ekran en aşağıda mı kontrolü 
                new_height = browser.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height
            
            count = len(users)
            print("Tweet sayısı: " + str(count))
            
            #Excel kayıt
            for i in range(0, count):
                ws.append([users[i],tweets[i]])
            wb.save("user-like-data.xlsx")
            browser.quit()
finally:
  print("Başarılı")




