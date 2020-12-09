import time                                            # İşlemler arası bekleme yapmak için kullanacağımız sleep() methodunun barındırıldığı kütüphane
import requests                                        # Scrapper için yardımcı kütüphane
from openpyxl import Workbook,load_workbook            # Excell'e veri kaydetmek içi kütüphane
from bs4 import BeautifulSoup                          # Proje konumuz olan kütüphane


try:                                                   # Scrapping'de beklenmedik hatalar olabiliyor, programın yarıda kalmaması için try finally bloğu kullanıyoruz
    Qut = False
    while Qut == False:                                # Kullanıcının programı tekrar başlatmadan çeşitli işlemler yapabilmesi için döngüye alıyoruz
    
        print("1:  Yaklaşan Filmler")                  # Seçime göre işlemler 
        print("2:  En Çok Oylanan Filmler")      
        print("3:  En Popüler Filmler")          
        print("4:  Bugün Doğan Ünlüler")         
        print("5:  En Popüler Ünlüler")          
        print("0: Çıkış")
        secim = int(input("İşlem Seçiniz.."+"\n"))
        
        wb = Workbook()                                # Excell çalışma dosyası tanımlıyoruz
        ws = wb.active                                 # Dosyamızı aktifleştiriyoruz
        ws.title = "Data"                              # Dosyamıza başlık verdik
    
        if secim < 0 or secim > 5:                     # Kullanıcı doğru bir seçim mi yapmış kontrolü
            print("Hatalı Seçim Yaptınız..")
            time.sleep(2)
        
        elif secim == 0:                               # Porgramdan çıkış işlemi
            Qut = True
            break
         
        elif secim == 1:                               # 1. işlemimiz yaklaşan filmlerin listelenmesi
            
            
            ws = wb.create_sheet("Yaklaşan Filmler")   # Yaklaşan Filmler adında excell sayfası oluşturuyoruz
    
            url = "https://www.imdb.com/calendar/?ref_=nv_mv_cal/"                                    # Filmlerin çekileceği url adresini tanımlıyoruz
            main = BeautifulSoup(requests.get(url).text, "html.parser").find("div", {"id": "main"})   # BeautifulSoup'la url adresimizi parse edip ardından içerisinde id'si "main" olan divi bulduruyoruz
            filmler = main.find_all("li")                                                             # Bulduğumuz divin içerisindeki tüm "li" taglarını listeye atıyoruz
            Filmler = []                                                                              # Filmlerin tutulacağı listeyi tanımlıyoruz
    
            ws.append([ "Filmadı", "Tarih", "Puan", "Sure", "Tur", "Yonetmen","Oyuncular","Dil" ])    # Excell tablo başlıklarını ekliyoruz
    
            for i in filmler:                                                                         # Az önce çekilip listeye atılan tüm "li" taglarımıza yani filmlerimizi döngüye sokuyoruz 
                Filmadı = i.a.text                                                                    # "li" tagı içerisindeki "a" tagının yazısını alıyoruz
                print(Filmadı)
    
                link = "https://www.imdb.com" + i.a.get("href")                                       # "li" tagı içerisindeki "a" tagının yönlendirme yaptığı url adresini alarak filmin detay sayfasına gidiyoruz
                detay = BeautifulSoup(requests.get(link).text, "html.parser")                         # Filmin detay sayfasını parse ediyoruz
                block = detay.find("div", {"class" : "title_block"})                                  # Detay sayfasında class'ı "title_block" olan divi buluyoruz
    
                #Çıkış tarihi########################
                Tarih = "None"                                                                        # Bazı filmlerin çıkış tarihleri eklenmiyor bu durumda programın hata vermemsi için girilmemiş tarihlerde "None" değeri döndürüyoruz
                if block.find("a", {"title" : "See more release dates"}):                             # Eğer filmin çıkış tarihi girilmiş isee..
                    tarih = block.find("a", {"title" : "See more release dates"})                     # title attribute'i "See more release dates" olan a tagını seçiyoruz
                    Tarih = tarih.text.strip()                                                        # Tarih metnini alınca içerisinde boşluk olabiliyor bu yüzden strip() metodu ile boşlukları kırpıyoruz
                    print(tarih.text.strip())
                
    
                #Puan################################
                Puan = "None"
                if block.find("div", {"class" : "ratingValue"}):                                      # Yukarıdaki Mantığın Benzeri
                    puan = block.find("div", {"class" : "ratingValue"}).text
                    print("Puan: "+puan.strip())
                    Puan = puan.strip()
                
                #Süre###############################
                Sure = "None"                                                                         # Yukarıdaki Mantığın Benzeri
                if block.find("time"): 
                    sure = block.find("time").text
                    print("Süre: "+sure.strip())
                Sure = sure.strip()
    
                #Tür################################
                Tur = "None"                                                                          # Yukarıdaki Mantığın Benzeri
                block1 = detay.find("div", {"class" : "subtext"})
                if block1.find("a"):
                    tur = block1.find("a").text
                    print("Tür: "+tur.strip()) 
                Tur = tur.strip()
    
                #Yonetmen##########################
                yonetmen = detay.find("div", {"class" : "plot_summary_wrapper"}).select("div > div:nth-of-type(2) > a:nth-of-type(1)") # istenilen div bulunduktan sonra, divin içerisindeki 2. div seçilir ve bu divin içindeki ilk a tagı seçilir
                for i in yonetmen:
                    print("Yonetmen: "+ i.text)
                    Yonetmen = i.text
                
                #Stars#############################
                stars = detay.find("div", {"class" : "plot_summary_wrapper"}).select("div > div:nth-of-type(4) > a")                   # Yukarıdaki Mantığın Benzeri
                stringstarts = ""
                for i in stars:
                    if i.text != "See full cast & crew":
                        stringstarts = stringstarts + i.text + ","          # Oyuncular 1 veya birden fazla olarak eklenebiliyor, eklenenlerin hepsini girmek için ayrıca for döngüsü kullanıyoruz
                print(stringstarts)
                Stars = stringstarts
    
                #Dil###############################
                dil = detay.find("div",{"id" : "titleDetails"}).select("div:nth-of-type(3) > a")                                       # Yukarıdaki Mantığın Aynısı
                for i in dil:
                    print("Dil: "+ i.text)
                    Dil = i.text
                ws.append([ Filmadı, Tarih, Puan, Sure, Tur, Yonetmen,Stars,Dil ])                                                     # Döngüdeki kayıt Excell'e kaydedilir
                wb.save("data.xlsx")                                                                                                   # tablo her ihtimale karşı her döngüde kaydedilir
        
        elif secim == 2:                                                                                                               # Diğer kodların yukarıdakilerden farklı bir yanı yok, benzerdir.
    
            Filmler = []
            Filmadi = []
            Filmyili = []
            Filmpuani = []
    
            ws = wb.create_sheet("En Çok Oylanan Filmler")
    
            ws.append([ "Film", "Yıl", "Puan"])
            url = "https://www.imdb.com/chart/top-english-movies"
            main = BeautifulSoup(requests.get(url).text, "html.parser").find("table")
            adiyili = main.find_all("td", {"class" : "titleColumn"})
            for i in adiyili:
                print("Film: "+ i.a.text)
                Filmadi.append(i.a.text.strip())
                print("Yıl: " + i.span.text)
                Filmyili.append(i.span.text.strip())
            
            puan = main.find_all("td",{"class" : "imdbRating"})
            for i in puan:
                Filmpuani.append(i.text.strip())
                print("Puan: "+ i.text.strip())
    
            count = len(Filmadi);
            print("Kayıtlı sayısı: " + str(count))
            
            for i in range(0,count-1):
                ws.append([ Filmadi[i],Filmyili[i],Filmpuani[i]])
                i = i + 1
            wb.save("data.xlsx")
        
        elif secim == 3:
    
            ws = wb.create_sheet("En Popüler Filmler")
            ws.append([ "Film", "Yıl", "Puan","Tür","Süre"])
    
            url = "https://www.imdb.com/chart/moviemeter/?ref_=nv_mv_mpm"
            main = BeautifulSoup(requests.get(url).text, "html.parser").find('tbody',{"class" : "lister-list"})
            filmler = main.find_all("tr")
    
            Filmler = []
    
            for i in filmler:
                print(i.find("td",{"class" : "titleColumn"}).a.text.strip())
                adi = i.find("td",{"class" : "titleColumn"}).a.text.strip()
    
                detaylink = i.find("td",{"class" : "titleColumn"}).a.get("href")
                detayurl = "https://www.imdb.com" + detaylink
                detay = BeautifulSoup(requests.get(detayurl).text, "html.parser")
                subtext = detay.find("div",{"class":"subtext"})
    
                sure = "none"
                if subtext.find("time"):
                    sure = subtext.find("time").text.strip()
    
                kategori = "none"
                if subtext.find("a"):
                    kategori = subtext.find("a").text.strip()
    
                print(i.find("td",{"class" : "titleColumn"}).span.text.strip())
                yılı = i.find("td",{"class" : "titleColumn"}).span.text.strip()
                puanı = "none"
                if i.find("td",{"class" : "imdbRating"}).strong:
                    print(i.find("td",{"class" : "imdbRating"}).strong.text.strip())
                    puanı = i.find("td",{"class" : "imdbRating"}).strong.text.strip()
                ws.append([ adi, yılı, puanı, kategori, sure, ])
                wb.save("data.xlsx")  
    
        elif secim == 4:
    
            ws = wb.create_sheet("Bugün Doğan Ünlüler")
    
            ws.append([ "Adı"])
            Stars =[]
            url = "https://www.imdb.com/chart/top-english-movies"
            main = BeautifulSoup(requests.get(url).text, "html.parser")
            todaylink = main.select("html > body > div:nth-of-type(1) > nav > div:nth-of-type(2) > aside > div > div:nth-of-type(2) > div > div:nth-of-type(4) > span > div > div > ul > a:nth-of-type(1)")
            for i in todaylink:
                print(i.get("href"))
                bugun = i.get("href")
    
            url2 = "https://www.imdb.com" + bugun
            print(url2)
            main2 = BeautifulSoup(requests.get(url2).text, "html.parser").find_all("h3",{"class" : "lister-item-header"})
    
            for i in main2:
                print(i.a.text.strip())
                Stars.append(i.a.text.strip())
            
            
            for i in Stars:
                ws.append([i])
                wb.save("data.xlsx")  
                     
        elif secim == 5:
    
            ws = wb.create_sheet("En Popüler Ünlüler")
            ws.append([ "Adı"])
            Stars = []
            url = "https://m.imdb.com/chart/starmeter/?ref_=nv_cel_brn"
            main = BeautifulSoup(requests.get(url).text, "html.parser").find("section",{"id" : "chart-content"})
            stars = main.find_all("h4")
    
            for i in stars:
                print(i.text.strip())
                Stars.append(i.text.strip())
            
            
            for i in Stars:
                ws.append([i])
                wb.save("data.xlsx")  
finally:
    print("başarılı")
  

